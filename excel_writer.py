"""Turn reconstructed tables into a plain, ready-to-use workbook.

The sheets look like ordinary spreadsheets: gridlines on, no fills, no
borders, one sheet per table. The only formatting applied is the kind that
carries meaning rather than decoration -- amounts become real Excel numbers so
they add up, dates become real dates, and a cell that spans the table in the
source document spans it here too.

Typing is deliberately conservative: a column is converted only when most of
its cells agree, and anything ambiguous is left as text so no value is ever
silently altered.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from table_engine import Document, Table

MIN_COL_WIDTH = 9
MAX_COL_WIDTH = 52
TYPE_CONFIDENCE = 0.7  # share of cells that must parse before a column is typed

CURRENCY_SYMBOLS = "₹$€£¥"
_CURRENCY_RE = re.compile(rf"[{CURRENCY_SYMBOLS}]|\b(?:INR|USD|EUR|GBP|AED|Rs\.?)\b", re.IGNORECASE)
_DRCR_RE = re.compile(r"[\s(\[]*\b(CR|DR)\b\.?[\s)\]]*$", re.IGNORECASE)
_PERCENT_RE = re.compile(r"%\s*$")

# Positive renders as Cr, negative as Dr -- Excel drops the sign in the negative
# section, so the cell reads exactly like the statement while still summing.
DRCR_FORMAT = '#,##0.00" Cr";#,##0.00" Dr";#,##0.00'


def split_drcr(text: str) -> Tuple[str, Optional[str]]:
    """Separate a trailing debit/credit marker from the amount before it."""
    raw = (text or "").strip()
    match = _DRCR_RE.search(raw)
    if not match:
        return raw, None
    return raw[: match.start()].strip(), match.group(1).upper()


# ---------------------------------------------------------------------------
# Value parsing
# ---------------------------------------------------------------------------


def parse_number(text: str) -> Optional[float]:
    """Parse an accounting-style amount, or return ``None``.

    Handles thousands separators, currency symbols and prefixes, percent signs
    and parenthesised negatives. Debit/credit markers are handled a level up in
    :func:`profile_column`, so strip them before calling this.
    """
    raw = (text or "").strip()
    if not raw:
        return None

    negative = raw.startswith("(") and raw.endswith(")")
    cleaned = _CURRENCY_RE.sub("", raw.strip("()")).replace("%", "")
    cleaned = cleaned.replace(",", "").replace(" ", "").strip()
    if cleaned.endswith("-"):  # trailing-minus convention
        negative, cleaned = True, cleaned[:-1].strip()

    if not re.fullmatch(r"[-+]?\d*\.?\d+", cleaned):
        return None

    try:
        value = float(cleaned)
    except ValueError:
        return None
    return -value if negative else value


_DATE_FORMATS = [
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%Y-%m-%d", "%Y/%m/%d",
    "%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y",
    "%b %d, %Y", "%B %d, %Y", "%b %d %Y",
    "%d/%m/%y", "%d-%m-%y", "%d-%b-%y", "%d %b %y",
]
_US_FORMATS = ["%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y"]
_DATE_SHAPE_RE = re.compile(r"^\s*\d{1,4}[/\-. ][A-Za-z0-9]{1,9}[/\-. ]\d{2,4}\s*$|^[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}$")


def parse_date(text: str, prefer_us: bool = False) -> Optional[datetime]:
    raw = (text or "").strip()
    if not raw or len(raw) > 24 or not _DATE_SHAPE_RE.match(raw):
        return None

    formats = (_US_FORMATS + _DATE_FORMATS) if prefer_us else (_DATE_FORMATS + _US_FORMATS)
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _detect_us_dates(values: Sequence[str]) -> bool:
    """Disambiguate d/m/y from m/d/y using the whole column."""
    day_first = month_first = 0
    for value in values:
        match = re.match(r"^\s*(\d{1,2})[/\-.](\d{1,2})[/\-.]\d{2,4}\s*$", value or "")
        if not match:
            continue
        first, second = int(match.group(1)), int(match.group(2))
        if first > 12 >= second:
            day_first += 1
        elif second > 12 >= first:
            month_first += 1
    return month_first > day_first


# ---------------------------------------------------------------------------
# Column typing
# ---------------------------------------------------------------------------


@dataclass
class ColumnSpec:
    kind: str = "text"            # text | number | integer | date | percent
    number_format: Optional[str] = None
    prefer_us_dates: bool = False
    width: float = MIN_COL_WIDTH
    wrap: bool = False


def _money_format(symbol: Optional[str], decimals: bool = True) -> str:
    body = "#,##0.00" if decimals else "#,##0"
    if symbol:
        return f'"{symbol}"\\ {body};[Red]-"{symbol}"\\ {body}'
    return f"{body};[Red]-{body}"


def profile_column(cells: Sequence[str]) -> ColumnSpec:
    values = [c.strip() for c in cells if c and c.strip()]
    spec = ColumnSpec()
    if not values:
        return spec

    # Ledger columns that carry a CR/DR marker become signed numbers (Dr
    # negative) with a format that prints the marker back. The cell then reads
    # like the statement and still adds up.
    marked = [v for v in values if split_drcr(v)[1]]
    if len(marked) / len(values) > 0.3:
        parsed = sum(1 for v in marked if parse_number(split_drcr(v)[0]) is not None)
        if parsed / len(marked) >= TYPE_CONFIDENCE:
            spec.kind = "drcr"
            spec.number_format = DRCR_FORMAT
            return spec
        spec.kind = "text"
        return spec

    prefer_us = _detect_us_dates(values)
    dates = sum(1 for v in values if parse_date(v, prefer_us) is not None)
    if dates / len(values) >= TYPE_CONFIDENCE:
        spec.kind = "date"
        spec.number_format = "dd-mmm-yyyy"
        spec.prefer_us_dates = prefer_us
        return spec

    numbers = [parse_number(v) for v in values]
    parsed = [n for n in numbers if n is not None]
    if len(parsed) / len(values) >= TYPE_CONFIDENCE:
        if sum(1 for v in values if _PERCENT_RE.search(v)) / len(values) > 0.5:
            spec.kind = "percent"
            spec.number_format = "0.00\\%"
            return spec

        symbol = _dominant_symbol(values)
        has_decimals = any("." in v for v in values)
        looks_like_id = not has_decimals and not symbol and all(
            abs(n) >= 1000 and float(n).is_integer() and "," not in v
            for n, v in zip(numbers, values) if n is not None
        )
        if looks_like_id:
            spec.kind = "text"  # invoice / account / cheque numbers
            return spec

        spec.kind = "integer" if not has_decimals else "number"
        spec.number_format = _money_format(symbol, decimals=has_decimals)
    return spec


def _dominant_symbol(values: Sequence[str]) -> Optional[str]:
    counts: Dict[str, int] = {}
    for value in values:
        for char in value:
            if char in CURRENCY_SYMBOLS:
                counts[char] = counts.get(char, 0) + 1
    if not counts:
        return None
    return max(counts, key=counts.get)


def coerce(value: str, spec: ColumnSpec) -> Any:
    text = (value or "").strip()
    if not text:
        return None
    if spec.kind == "date":
        parsed = parse_date(text, spec.prefer_us_dates)
        return parsed if parsed is None else parsed.date()
    if spec.kind == "drcr":
        amount, marker = split_drcr(text)
        number = parse_number(amount)
        if number is None:
            return text
        return -abs(number) if marker == "DR" else number
    if spec.kind in {"number", "integer", "percent"}:
        number = parse_number(text)
        return text if number is None else number
    return text


# ---------------------------------------------------------------------------
# Header hygiene
# ---------------------------------------------------------------------------


def _drop_unreadable_glyphs(name: str) -> str:
    """Tidy a heading whose currency symbol did not survive the PDF's font.

    Plenty of templates write ``Amount(₹)`` with the rupee sign mapped into a
    private-use slot, so it decodes to a blank, a replacement mark or a stray
    middle dot and the heading arrives as ``Amount( )``. Bracketed groups that
    carry no readable character are dropped; a symbol that *did* decode, and
    anything with letters or digits in it, is left exactly as it is.
    """

    def replace(match: re.Match) -> str:
        inner = match.group(1).strip()
        if not inner:
            return ""
        if any(c.isalnum() or c in CURRENCY_SYMBOLS for c in inner):
            return match.group(0)
        return ""

    return re.sub(r"\(([^()]{0,4})\)|\[([^\[\]]{0,4})\]", replace, name).strip(" .:-")


def clean_headers(header: Sequence[str], n_cols: int) -> List[str]:
    names: List[str] = []
    seen: Dict[str, int] = {}
    for i in range(n_cols):
        raw = " ".join((header[i] if i < len(header) else "").split())
        raw = raw.replace("\n", " ").strip(" .:-")
        raw = _drop_unreadable_glyphs(raw.replace("\xa0", " "))
        name = raw if raw else f"Column {i + 1}"
        key = name.lower()
        if key in seen:
            seen[key] += 1
            name = f"{name} ({seen[key]})"
        else:
            seen[key] = 1
        names.append(name[:120])
    return names


def _sheet_name(candidate: str, used: set) -> str:
    name = re.sub(r"[\\/*?:\[\]]", " ", candidate or "Table").strip() or "Table"
    name = " ".join(name.split())[:31] or "Table"
    base = name
    counter = 2
    while name.lower() in used:
        suffix = f" {counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(name.lower())
    return name


# ---------------------------------------------------------------------------
# Sheet rendering
# ---------------------------------------------------------------------------


def _write_table(ws: Worksheet, table: Table) -> None:
    """Write one table as a plain worksheet: header row, then data.

    Deliberately unstyled beyond a bold header -- gridlines stay on and the
    sheet looks like an ordinary spreadsheet. The only formatting applied is
    the kind that carries meaning: number and date formats, so amounts add up,
    and merges for cells that span the table in the source document.
    """
    n_cols = table.n_cols
    header = clean_headers(table.header, n_cols) if table.has_header else [
        f"Column {i + 1}" for i in range(n_cols)
    ]
    body = [list(row) + [""] * (n_cols - len(row)) for row in table.body]

    spans = spanning_rows(body, n_cols)
    # Spanning captions are not part of any column, so they must not influence
    # how that column is typed.
    typed_body = [row for i, row in enumerate(body) if i not in spans]
    specs = [profile_column([row[c] for row in typed_body]) for c in range(n_cols)]

    header_row = 1
    first_data_row = 2
    ws.freeze_panes = "A2"

    for col, name in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True)

    for offset, row_values in enumerate(body):
        excel_row = first_data_row + offset

        if offset in spans:
            _write_spanning_row(ws, excel_row, n_cols, row_values[spans[offset]])
            continue

        for col in range(1, n_cols + 1):
            spec = specs[col - 1]
            value = coerce(row_values[col - 1], spec)
            cell = ws.cell(row=excel_row, column=col, value=value)

            if isinstance(value, (int, float)) and spec.number_format:
                cell.number_format = spec.number_format
            elif isinstance(value, (date, datetime)):
                cell.number_format = spec.number_format or "dd-mmm-yyyy"
            elif isinstance(value, str) and "\n" in value:
                # A bordered cell holding several lines keeps its line breaks,
                # but Excel only honours them when wrapping is on.
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if body:
        last_row = first_data_row + len(body) - 1
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"

    # Merged captions are excluded so one long banner cannot widen a column.
    _autosize(ws, header, typed_body or body, specs, n_cols)


def _write_spanning_row(ws: Worksheet, row: int, n_cols: int, text: str) -> None:
    """Render a caption as one merged cell running the width of the table."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    anchor = ws.cell(row=row, column=1, value=text)
    anchor.font = Font(bold=True)


def spanning_rows(body: Sequence[Sequence[str]], n_cols: int) -> Dict[int, int]:
    """Find rows that carry one value across the whole table width.

    Section captions inside a table ("Part B - Deductions") and single-cell
    banners occupy one column here but span every column in the PDF. Genuine
    wrapped descriptions have already been folded into their parent row by the
    extractor, so what is left really is meant to span.

    Returns ``{row index: column index holding the text}``.
    """
    if n_cols < 3:
        return {}
    spans: Dict[int, int] = {}
    for i, row in enumerate(body):
        filled = [c for c, cell in enumerate(row) if cell.strip()]
        if len(filled) == 1 and not _looks_like_amount(row[filled[0]]):
            spans[i] = filled[0]
    return spans


def _looks_like_amount(text: str) -> bool:
    return parse_number(split_drcr(text)[0]) is not None


def _autosize(
    ws: Worksheet,
    header: Sequence[str],
    body: Sequence[Sequence[str]],
    specs: Sequence[ColumnSpec],
    n_cols: int,
) -> None:
    for col in range(n_cols):
        lengths = [len(str(header[col]))]
        lengths.extend(len(str(row[col] or "")) for row in body)
        widest = max(lengths) if lengths else MIN_COL_WIDTH
        typical = sorted(lengths)[int(len(lengths) * 0.9)] if len(lengths) > 4 else widest

        if specs[col].kind == "text" and widest > MAX_COL_WIDTH:
            width = MAX_COL_WIDTH
            specs[col].wrap = True
        else:
            width = min(MAX_COL_WIDTH, max(MIN_COL_WIDTH, typical + 3))
        ws.column_dimensions[get_column_letter(col + 1)].width = width

    if any(spec.wrap for spec in specs):
        for row in range(4, 4 + len(body)):
            ws.row_dimensions[row].height = None  # let Excel grow wrapped rows


def _write_key_values(ws: Worksheet, pairs) -> None:
    """A plain two-column sheet of the labels found around the tables."""
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 58
    for col, name in enumerate(("Field", "Value"), start=1):
        ws.cell(row=1, column=col, value=name).font = Font(bold=True)
    for i, (key, value) in enumerate(pairs, start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def build_workbook(document: Document, filename: str = "document.pdf") -> bytes:
    """One sheet per table, plus the document's label/value fields if any."""
    wb = Workbook()
    wb.remove(wb.active)

    used: set = set()
    for i, table in enumerate(document.tables, start=1):
        name = _sheet_name(table.title or f"Table {i}", used)
        _write_table(wb.create_sheet(name), table)

    if document.key_values:
        _write_key_values(wb.create_sheet(_sheet_name("Document Details", used)),
                          document.key_values)
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
